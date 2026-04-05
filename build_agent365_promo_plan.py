#!/usr/bin/env python3
"""
Промо-план Agent #365 (ИП Чепкасова Н. А.) — Красноярск.
Данные: январь + февраль + март 2026 (orderTable CSV).

RFM: квантильные баллы R, F, M (1–5). Сегменты в отчёте — классическая карта
(Champions, Loyal, …) как в AquaBot_Samara / единой схеме для владельцев.

Структура Excel выровнена по H2O_Saransk_Promo_Plan.xlsx (заголовки, блоки, листы).
"""
from __future__ import annotations

import json
import re
import urllib.request
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

YANDEX_PHONE = "133133133133"
TBANK_PHONE = "71119999991"
AGENT_PATTERN = re.compile(r"Agent\s*#\s*365", re.I)
TIME_WINDOW_MIN = 30

WD_RU = (
    "Понедельник",
    "Вторник",
    "Среда",
    "Четверг",
    "Пятница",
    "Суббота",
    "Воскресенье",
)


def load_order_table(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, sep=";", encoding="utf-8-sig", dtype=str)
    num_cols = [
        "Поступило на бокс",
        "Оплачено деньгами",
        "Оплачено бонусами",
        "Начислено кешбека",
    ]
    for col in num_cols:
        if col in df.columns:
            s = (
                df[col]
                .astype(str)
                .str.replace("\u00a0", " ", regex=False)
                .str.replace(" ", "", regex=False)
                .str.replace(",", ".", regex=False)
            )
            df[col] = pd.to_numeric(s, errors="coerce").fillna(0.0)
    df["Дата оплаты"] = pd.to_datetime(df["Дата оплаты"], errors="coerce")
    df["date"] = df["Дата оплаты"].dt.date
    df["total"] = df.get("Оплачено деньгами", 0) + df.get("Оплачено бонусами", 0)
    return df


def normalize_phone(phone: str) -> str:
    if pd.isna(phone) or str(phone).strip() == "":
        return ""
    s = str(phone).strip().replace(",", ".")
    try:
        return str(int(float(s)))
    except (ValueError, OverflowError):
        return "".join(filter(str.isdigit, s))


def categorize_phone(phone: str) -> str:
    p = normalize_phone(phone)
    if p == YANDEX_PHONE:
        return "Яндекс"
    if p == TBANK_PHONE:
        return "Т-Банк"
    return "Лейка"


def group_transactions_to_visits(df: pd.DataFrame) -> pd.DataFrame:
    empty = pd.DataFrame(
        columns=[
            "visit_id",
            "visit_total",
            "Оплачено деньгами",
            "Оплачено бонусами",
            "Поступило на бокс",
            "Телефон",
            "Дата оплаты",
        ]
    )
    if df.empty or "Дата оплаты" not in df.columns or "Телефон" not in df.columns:
        return empty
    w = df.dropna(subset=["Дата оплаты", "Телефон"]).copy()
    if w.empty:
        return empty
    w = w.sort_values(["Телефон", "Дата оплаты"])
    prev = w.groupby("Телефон")["Дата оплаты"].shift(1)
    gap = (w["Дата оплаты"] - prev).dt.total_seconds() / 60
    w["visit_id"] = (gap.isna() | (gap > TIME_WINDOW_MIN)).cumsum()
    visits = (
        w.groupby("visit_id")
        .agg(
            total=("total", "sum"),
            **{
                "Оплачено деньгами": ("Оплачено деньгами", "sum"),
                "Оплачено бонусами": ("Оплачено бонусами", "sum"),
                "Поступило на бокс": ("Поступило на бокс", "sum"),
                "Телефон": ("Телефон", "first"),
                "Дата оплаты": ("Дата оплаты", "first"),
            },
        )
        .reset_index()
        .rename(columns={"total": "visit_total"})
    )
    return visits


def quintile_score_1_to_5(s: pd.Series, *, higher_is_better: bool) -> pd.Series:
    """
    Квантильные баллы 1–5 по рангу (эквивалент разбиения на 5 групп по численности).
    higher_is_better=False: меньшее значение (например дни с последнего визита) → балл 5.
    """
    rank = s.rank(method="first")
    if not higher_is_better:
        rank = rank.max() + 1 - rank
    n = len(s)
    if n == 0:
        return pd.Series(dtype=int)
    pct = (rank - 0.5) / n
    scores = np.clip(np.ceil(pct * 5).astype(int), 1, 5)
    return pd.Series(scores.values, index=s.index, dtype=int)


def rfm_code_label(r: int, f: int, m: int) -> str:
    return f"{r}-{f}-{m}"


# Порядок строк как в AquaBot_Samara_PREDECTIVE.xlsx (лист «Бюджет промо» / «RFM сегменты»)
SEGMENT_ORDER: tuple[str, ...] = (
    "Champions",
    "Loyal",
    "Promising",
    "Need Attention",
    "At Risk",
    "About to Sleep",
    "Hibernating",
)

# Русские названия инициатив — как в Samara
INITIATIVE_RU: dict[str, str] = {
    "Champions": "Удержание лучших",
    "Loyal": "Рост частоты",
    "Promising": "Конверсия новых",
    "Need Attention": "Лёгкая реактивация",
    "At Risk": "Срочная реактивация",
    "About to Sleep": "Реактивация уходящих",
    "Hibernating": "Возврат потерянных",
}

# Типичный R в сегменте, если клиентов 0 (для шаблона строк)
_DEFAULT_R_BY_SEGMENT: dict[str, int] = {
    "Champions": 5,
    "Loyal": 4,
    "Promising": 5,
    "Need Attention": 3,
    "At Risk": 2,
    "About to Sleep": 2,
    "Hibernating": 1,
}


def marketing_segment(r: int, f: int, m: int) -> str:
    """
    Классическая 7-сегментная карта по квантилям R,F,M (5 = лучшая свежесть).
    Правила согласованы с отчётами AquaBot Samara / сетевыми предиктивками.
    """
    if r >= 4 and f >= 4 and m >= 4:
        return "Champions"
    if r >= 4 and f >= 4:
        return "Loyal"
    if r >= 4 and f <= 2:
        return "Promising"
    if r >= 4:
        return "Loyal"
    if r == 3 and f >= 3 and m >= 3:
        return "Need Attention"
    if r == 2 and f >= 3 and m >= 3:
        return "At Risk"
    if r == 2:
        return "About to Sleep"
    if r == 3:
        return "About to Sleep"
    if r == 1:
        return "Hibernating"
    return "Hibernating"


def promo_bundle_for_segment(segment: str) -> tuple[str, str, str, str]:
    """Промо в духе AquaBot_Samara. Первое поле — код в Excel; для «Бонусы на счёт» это код кампании начисления, не купон на скидку."""
    bundles: dict[str, tuple[str, str, str, str]] = {
        "Champions": (
            "VIP50KR365",
            "Бонусы на счёт",
            "50₽ бонус на счёт",
            "1 — Высший (удержание VIP)",
        ),
        "Loyal": (
            "PLUS30KR365",
            "Бонусы на счёт",
            "30₽ бонус на счёт",
            "2 — Высокий (рост частоты)",
        ),
        "Promising": (
            "START50KR365",
            "Бонусы на счёт",
            "50₽ бонус на счёт",
            "3 — Средний (конверсия)",
        ),
        "Need Attention": (
            "MISS100KR365",
            "100% оплата промокодом",
            "Бесплатная мойка до 100₽",
            "4 — Высокий (реактивация)",
        ),
        "At Risk": (
            "BACK150KR365",
            "100% оплата промокодом",
            "Бесплатная мойка до 150₽",
            "5 — Очень высокий (срочная реактивация)",
        ),
        "About to Sleep": (
            "WAKE100KR365",
            "100% оплата промокодом",
            "Бесплатная мойка до 100₽",
            "6 — Средний (реактивация)",
        ),
        "Hibernating": (
            "SPRING200KR365",
            "100% оплата промокодом",
            "Бесплатная мойка до 200₽",
            "7 — Низкий (холодная реактивация)",
        ),
    }
    return bundles[segment]


# Тексты push для сегментов «Бонусы на счёт»: суммы 50/30/50 ₽ — это начисления на бонусный счёт Лейки, не скидка по промокоду.
_PUSH_BONUS_ON_ACCOUNT: dict[str, str] = {
    "Champions": (
        "Спасибо за выбор мойки на Светлогорской! "
        "Вам начислен VIP-бонус 50₽ на бонусный счёт. "
        "Спасибо за вашу лояльность."
    ),
    "Loyal": (
        "Спасибо за выбор мойки на Светлогорской! "
        "Вам начислен бонус 30₽ на бонусный счёт. "
        "Спасибо за вашу лояльность."
    ),
    "Promising": (
        "Спасибо за выбор мойки на Светлогорской! "
        "Вам начислен бонус 50₽ на бонусный счёт. "
        "Будем рады снова увидеть вас на мойке!"
    ),
}


def push_message_text(segment: str, pcode: str, summa: str, mech: str, rec: str) -> str:
    """Push пачками, без имени. Для бонусов на счёт — без «промокод», для механики 100% — с кодом активации."""
    if mech == "Бонусы на счёт":
        return _PUSH_BONUS_ON_ACCOUNT[segment]
    tail = summa[:1].lower() + summa[1:] if summa else ""
    return (
        f"Спасибо за выбор мойки на Светлогорской! "
        f"Активируйте промокод {pcode} в приложении — {tail}. "
        f"({rec})"
    )


def rep_r_for_segment(clients: pd.DataFrame, segment: str) -> int:
    """Репрезентативный квантиль R для оценки бюджета по сегменту."""
    sub = clients[clients["Сегмент"] == segment]["R"]
    if len(sub) == 0:
        return _DEFAULT_R_BY_SEGMENT[segment]
    return int(round(float(sub.median())))


def budget_estimate_for_r(r: int, n: int) -> tuple[int, str, str]:
    """Оценка бюджета и периодичности по доминирующему R в ячейке."""
    if r >= 4:
        return n * 50, "Ежемесячно", "~оценка удержания"
    if r == 3:
        return n * 35, "Ежемесячно", "~оценка частоты"
    if r == 2:
        return n * 100, "Однократно", "~оценка реактивации"
    return n * 200, "Однократно", "~оценка возврата"


def fetch_weather_krasnoyarsk(d0, d1) -> pd.DataFrame:
    lat, lon = 56.0184, 92.8672
    url = (
        "https://archive-api.open-meteo.com/v1/archive?"
        f"latitude={lat}&longitude={lon}"
        "&start_date={}&end_date={}"
        "&daily=temperature_2m_max,temperature_2m_min,precipitation_sum,"
        "snowfall_sum,rain_sum,weathercode"
        "&timezone=Asia%2FKrasnoyarsk"
    ).format(d0.isoformat(), d1.isoformat())
    with urllib.request.urlopen(url, timeout=60) as resp:
        data = json.loads(resp.read().decode())
    daily = data["daily"]
    return pd.DataFrame(
        {
            "date": pd.to_datetime(pd.Series(daily["time"])).dt.date,
            "temp_max": daily["temperature_2m_max"],
            "temp_min": daily["temperature_2m_min"],
            "precipitation": daily["precipitation_sum"],
            "snowfall": daily["snowfall_sum"],
            "rain": daily["rain_sum"],
            "weathercode": daily["weathercode"],
        }
    )


def append_df(ws, df: pd.DataFrame, start_row: int, start_col: int = 1) -> int:
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=0):
        for c_idx, val in enumerate(row, start=0):
            ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=val)
    return start_row + len(df) + 1


def apply_visual_style_saransk(
    wb: Workbook,
    ref_path: Path,
    *,
    rfm_next: int,
    rfm2_next: int,
) -> None:
    """
    Ширины колонок с эталона H2O_Saransk_Promo_Plan.xlsx, синие шапки, границы,
    зелёная/оранжевая подсветка K–L на листе «Клиенты для промо», закрепление областей.
    """
    thin = Side(style="thin", color="FFD0D0D0")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="FF4472C4")
    hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=11, bold=True)
    fill_bonus = PatternFill("solid", fgColor="FFE2EFDA")
    fill_promo = PatternFill("solid", fgColor="FFFCE4D6")

    if ref_path.exists():
        ref = load_workbook(ref_path)
        try:
            for name in wb.sheetnames:
                if name not in ref.sheetnames:
                    continue
                tw, rw = wb[name], ref[name]
                for col, dim in rw.column_dimensions.items():
                    if dim.width:
                        tw.column_dimensions[col].width = dim.width
        finally:
            ref.close()
    else:
        # резервные ширины, если эталона нет рядом
        fallback = {
            "Обзор и модель": {"A": 55, "B": 36, "S": 23},
            "RFM сегменты": {"A": 46, "K": 55},
            "Клиенты для промо": {"A": 55, "N": 55},
            "Погода и загрузка": {"A": 42, "M": 40},
            "Когорты retention": {"A": 45},
            "Бюджет промо": {"B": 55, "K": 55},
        }
        for sn, cols in fallback.items():
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            for letter, w in cols.items():
                ws.column_dimensions[letter].width = w

    def paint_header(ws, row: int, max_col: int) -> None:
        for c in range(1, max_col + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.border = grid
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def paint_grid(ws, r1: int, r2: int, c1: int, c2: int) -> None:
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                ws.cell(row=r, column=c).border = grid

    # --- Обзор и модель ---
    ws0 = wb["Обзор и модель"]
    paint_header(ws0, 5, 6)
    paint_header(ws0, 14, 19)
    paint_grid(ws0, 6, 6, 1, 6)
    paint_grid(ws0, 15, 15, 1, 19)
    paint_grid(ws0, 21, 21, 1, 19)
    ws0.cell(4, 1).font = title_font
    ws0.cell(13, 1).font = title_font
    for r in (1, 2):
        ws0.cell(r, 1).alignment = Alignment(wrap_text=True)

    # --- RFM сегменты ---
    ws1 = wb["RFM сегменты"]
    paint_header(ws1, 4, 11)
    paint_grid(ws1, 5, rfm_next - 1, 1, 11)
    ws1.cell(rfm_next, 1).font = title_font
    hr2 = rfm_next + 1
    paint_header(ws1, hr2, 6)
    paint_grid(ws1, hr2 + 1, rfm2_next - 1, 1, 6)
    ws1.freeze_panes = "A5"

    # --- Клиенты для промо (K,L как в Саранске: бонусы зелёные, 100% оранжевые) ---
    ws2 = wb["Клиенты для промо"]
    paint_header(ws2, 4, 15)
    max_r = ws2.max_row
    for r in range(5, max_r + 1):
        mech = ws2.cell(r, 11).value
        m = str(mech or "")
        row_fill = None
        if "Бонусы" in m:
            row_fill = fill_bonus
        elif "100%" in m or "промокод" in m.lower():
            row_fill = fill_promo
        for c in range(1, 16):
            cell = ws2.cell(r, c)
            cell.border = grid
            if row_fill is not None and c in (11, 12):
                cell.fill = row_fill
    ws2.freeze_panes = "A5"

    # --- Погода и загрузка ---
    ws3 = wb["Погода и загрузка"]
    paint_header(ws3, 3, 13)
    paint_grid(ws3, 4, ws3.max_row, 1, 13)
    ws3.freeze_panes = "A4"

    # --- Когорты retention ---
    ws4 = wb["Когорты retention"]
    paint_header(ws4, 4, 3)
    paint_grid(ws4, 5, 11, 1, 3)
    ws4.cell(13, 1).font = title_font
    paint_header(ws4, 14, 7)
    paint_grid(ws4, 15, 15, 1, 7)

    # --- Бюджет промо ---
    ws5 = wb["Бюджет промо"]
    paint_header(ws5, 4, 11)
    for r in range(5, ws5.max_row + 1):
        for c in range(1, 12):
            if ws5.cell(r, c).value is not None or c <= 2:
                ws5.cell(r, c).border = grid


def main() -> None:
    root = Path(__file__).resolve().parent
    files = [
        root / "leyka_january2026.csv",
        root / "february2026.csv",
        root / "leyka_march_2026.csv",
    ]
    for f in files:
        if not f.exists():
            raise SystemExit(f"Нет файла: {f}")

    raw = pd.concat([load_order_table(f) for f in files], ignore_index=True)
    raw["partner_category"] = raw["Телефон"].apply(categorize_phone)

    agent_mask = raw["Партнёр"].astype(str).apply(lambda s: bool(AGENT_PATTERN.search(s)))
    agent_df = raw[agent_mask].copy()
    wash = agent_df[agent_df["Тип оплаты"].astype(str).str.strip() == "Мойка автомобиля"].copy()
    leyka = wash[wash["partner_category"] == "Лейка"].copy()

    analysis_end = pd.Timestamp(leyka["Дата оплаты"].max()).normalize()
    if pd.isna(analysis_end):
        raise SystemExit("Пустые даты")

    visits = group_transactions_to_visits(leyka)
    last_visit = visits.groupby("Телефон")["Дата оплаты"].max()
    vc = visits.groupby("Телефон").size()
    ltv = visits.groupby("Телефон")["visit_total"].sum()
    clients = pd.DataFrame(
        {
            "Телефон": ltv.index,
            "LTV": ltv.values,
            "визитов": vc.reindex(ltv.index).values,
            "последний_визит": last_visit.reindex(ltv.index).values,
        }
    ).reset_index(drop=True)
    clients["recency_дней"] = (
        analysis_end - pd.to_datetime(clients["последний_визит"])
    ).dt.days.astype(int)

    clients["R"] = quintile_score_1_to_5(clients["recency_дней"], higher_is_better=False)
    clients["F"] = quintile_score_1_to_5(clients["визитов"].astype(float), higher_is_better=True)
    clients["M"] = quintile_score_1_to_5(clients["LTV"].astype(float), higher_is_better=True)
    clients["RFM_код"] = [
        rfm_code_label(int(r), int(f), int(m))
        for r, f, m in zip(clients["R"], clients["F"], clients["M"])
    ]
    clients["Сегмент"] = [
        marketing_segment(int(r), int(f), int(m))
        for r, f, m in zip(clients["R"], clients["F"], clients["M"])
    ]

    last_tx = (
        leyka.sort_values("Дата оплаты")
        .groupby("Телефон")
        .tail(1)[["Телефон", "Клиент"]]
        .rename(columns={"Клиент": "Имя клиента"})
    )
    clients = clients.merge(last_tx, on="Телефон", how="left")

    total_rev = clients["LTV"].sum()
    n_clients = len(clients)

    # --- Сводка RFM по маркетинговым сегментам (как AquaBot Samara) ---
    seg_counts = clients["Сегмент"].value_counts()
    seg_rev = clients.groupby("Сегмент", observed=False)["LTV"].sum()
    rfm_rows: list[dict] = []
    for seg in SEGMENT_ORDER:
        n = int(seg_counts.get(seg, 0))
        rev = float(seg_rev.get(seg, 0.0)) if n else 0.0
        sub = clients[clients["Сегмент"] == seg]
        sv = float(sub["визитов"].mean()) if n else 0.0
        sd = float(sub["recency_дней"].mean()) if n else 0.0
        pcode, mech, sum_txt, rec = promo_bundle_for_segment(seg)
        rfm_rows.append(
            {
                "Сегмент": seg,
                "Клиентов": n,
                "Доля базы": (n / n_clients) if n_clients else 0.0,
                "Выручка": rev,
                "Доля выручки": (rev / total_rev) if total_rev else 0.0,
                "Ср. визитов": sv,
                "Ср. дней с посл. визита": sd,
                "Промокод": pcode,
                "Механика промокода": mech,
                "Сумма": sum_txt,
                "Рекомендуемое действие": rec,
            }
        )
    rfm_agg = pd.DataFrame(rfm_rows)

    # Когорты
    leyka["month"] = pd.to_datetime(leyka["date"]).dt.to_period("M").astype(str)
    jan_phones = set(leyka[leyka["month"] == "2026-01"]["Телефон"].unique())
    feb_phones = set(leyka[leyka["month"] == "2026-02"]["Телефон"].unique())
    mar_phones = set(leyka[leyka["month"] == "2026-03"]["Телефон"].unique())
    ret_j_f = len(jan_phones & feb_phones) / len(jan_phones) if jan_phones else 0
    ret_f_m = len(feb_phones & mar_phones) / len(feb_phones) if feb_phones else 0
    core = jan_phones & feb_phones & mar_phones

    visits["month"] = visits["Дата оплаты"].dt.to_period("M").astype(str)
    obj_name = wash["Автомойка"].iloc[0] if len(wash) else ""
    addr = wash["Адрес"].iloc[0] if len(wash) else ""

    def month_stats(m: str):
        lv = leyka[leyka["month"] == m]
        vv = visits[visits["month"] == m]
        return {
            "tx": len(lv),
            "vis": len(vv),
            "cli": lv["Телефон"].nunique(),
            "rev": lv["total"].sum(),
        }

    mjan, mfeb, mmar = [month_stats(x) for x in ("2026-01", "2026-02", "2026-03")]
    total_tx = mjan["tx"] + mfeb["tx"] + mmar["tx"]
    total_vis = mjan["vis"] + mfeb["vis"] + mmar["vis"]
    total_cli = len(clients)
    total_rev_all = float(leyka["total"].sum())
    avg_check_visit = total_rev_all / len(visits) if len(visits) else 0

    # Погода + загрузка (как в Саранске: порядок колонок)
    d0, d1 = wash["date"].min(), wash["date"].max()
    weather = fetch_weather_krasnoyarsk(d0, d1)
    daily_vis = visits.copy()
    daily_vis["d"] = daily_vis["Дата оплаты"].dt.date
    daily_v = daily_vis.groupby("d").agg(визитов=("visit_id", "count"), клиентов=("Телефон", "nunique"))
    daily_tx = leyka.groupby(leyka["Дата оплаты"].dt.date).agg(Выручка=("total", "sum"), Транзакций=("total", "size"))
    load_w = weather.merge(daily_v, left_on="date", right_index=True, how="left")
    load_w = load_w.merge(daily_tx, left_on="date", right_index=True, how="left")
    load_w["визитов"] = load_w["визитов"].fillna(0).astype(int)
    load_w["клиентов"] = load_w["клиентов"].fillna(0).astype(int)
    load_w["Транзакций"] = load_w["Транзакций"].fillna(0).astype(int)
    load_w["Выручка"] = load_w["Выручка"].fillna(0.0)
    load_w["Транз/визит"] = np.where(
        load_w["визитов"] > 0, load_w["Транзакций"] / load_w["визитов"], 0.0
    )
    load_w["T средн °C"] = (load_w["temp_max"] + load_w["temp_min"]) / 2
    _dt = pd.to_datetime(load_w["date"])
    load_w["День недели"] = [_dt.iloc[i].dayofweek for i in range(len(load_w))]
    load_w["День недели"] = load_w["День недели"].apply(lambda i: WD_RU[int(i)])
    load_w["Выходной"] = _dt.dt.weekday >= 5
    load_w["Выходной"] = load_w["Выходной"].map({True: "Да", False: "Нет"})

    # Клиенты для промо (колонки 1:1 как в H2O_Saransk_Promo_Plan.xlsx; сегмент — как в Samara)
    promo_rows = []
    for _, row in clients.iterrows():
        seg = str(row["Сегмент"])
        pcode, mech, summa, rec = promo_bundle_for_segment(seg)
        name = str(row.get("Имя клиента") or "").strip()
        # Push уходит пачками по спискам — без обращения по имени (не персонализируем «Андрей», «Иван» и т.п.).
        promo_rows.append(
            {
                "Телефон": row["Телефон"],
                "Имя клиента": name,
                "Объекты посещения": f"{obj_name}, {addr}",
                "Сегмент": seg,
                "LTV (общая сумма оплат)": row["LTV"],
                "Количество визитов": int(row["визитов"]),
                "Количество транзакций": int(leyka[leyka["Телефон"] == row["Телефон"]].shape[0]),
                "Дата последнего визита": pd.to_datetime(row["последний_визит"]).strftime("%d.%m.%Y"),
                "Дней с последнего визита": int(row["recency_дней"]),
                "Промокод": pcode,
                "Механика": mech,
                "Сумма промо": summa,
                "Канал": "Push",
                "Текст сообщения": push_message_text(seg, pcode, summa, mech, rec),
                "Приоритет": f"{seg} — {INITIATIVE_RU.get(seg, seg)}",
            }
        )
    seg_rank = {s: i for i, s in enumerate(SEGMENT_ORDER)}
    promo_df = pd.DataFrame(promo_rows)
    promo_df["_sr"] = promo_df["Сегмент"].map(seg_rank)
    promo_df = promo_df.sort_values(["_sr", "LTV (общая сумма оплат)"], ascending=[True, False]).drop(
        columns=["_sr"]
    )

    # Бюджет: 7 маркетинговых инициатив + ИТОГО (как AquaBot_Samara_PREDECTIVE.xlsx)
    budget_rows = []
    for seg in SEGMENT_ORDER:
        gr = rfm_agg[rfm_agg["Сегмент"] == seg].iloc[0]
        n = int(gr["Клиентов"])
        r_rep = rep_r_for_segment(clients, seg)
        pcode, mech, sum_txt, _ = promo_bundle_for_segment(seg)
        bud, period, roi_hint = budget_estimate_for_r(r_rep, n) if n else (0, "—", "—")
        budget_rows.append(
            {
                "Инициатива": INITIATIVE_RU[seg],
                "Сегмент": seg,
                "Промокод": pcode,
                "Механика": mech,
                "Сумма": sum_txt,
                "Клиентов": n,
                "Бюджет": int(bud),
                "Периодичность": period,
                "Ожидаемый доп. доход/мес": roi_hint,
                "Окупаемость": "—",
                "Текст сообщения": (
                    f"Начисление бонусов на счёт (код кампании {pcode}) — см. «Клиенты для промо»"
                    if mech == "Бонусы на счёт"
                    else f"Промокод на оплату мойки {pcode} — см. «Клиенты для промо»"
                ),
            }
        )
    total_bud = sum(b["Бюджет"] for b in budget_rows)
    budget_rows.append(
        {
            "Инициатива": "ИТОГО",
            "Сегмент": None,
            "Промокод": None,
            "Механика": None,
            "Сумма": None,
            "Клиентов": n_clients,
            "Бюджет": total_bud,
            "Периодичность": None,
            "Ожидаемый доп. доход/мес": "~оценка",
            "Окупаемость": "< 2–3 мес (оценка)",
            "Текст сообщения": None,
        }
    )
    budget_df = pd.DataFrame(budget_rows)

    out_xlsx = root / "Agent365_Chepkasov_Promo_Plan.xlsx"
    out_md = root / "AGENT_365_CHEPKASOV_PREDICTIVE.md"

    days_count = (pd.Timestamp(d1) - pd.Timestamp(d0)).days + 1
    city = str(addr).split(",")[0].strip() if addr else "Красноярск"
    jan_u = len(set(leyka[leyka["month"] == "2026-01"]["Телефон"]))
    feb_u = len(set(leyka[leyka["month"] == "2026-02"]["Телефон"]))
    mar_u = len(set(leyka[leyka["month"] == "2026-03"]["Телефон"]))

    # Таблица RFM для листа (колонки как в Samara / Саранске)
    rfm_sheet_a = rfm_agg[
        [
            "Сегмент",
            "Клиентов",
            "Доля базы",
            "Выручка",
            "Доля выручки",
            "Ср. визитов",
            "Ср. дней с посл. визита",
            "Промокод",
            "Механика промокода",
            "Сумма",
            "Рекомендуемое действие",
        ]
    ]
    rfm_sheet_b = rfm_agg[["Сегмент", "Клиентов", "Доля базы", "Выручка", "Ср. визитов", "Ср. дней с посл. визита"]].rename(
        columns={"Доля базы": "Доля"}
    )

    wb = Workbook()

    # ----- Обзор и модель (строки как в H2O_Saransk_Promo_Plan) -----
    ws0 = wb.active
    ws0.title = "Обзор и модель"
    ws0.cell(1, 1, "Предиктивная аналитика — Agent #365 (ИП Чепкасова Н. А.) г. Красноярск")
    ws0.cell(
        2,
        1,
        f"Данные: январь — март 2026 ({days_count} дней), 1 объект, клиентов Лейки для RFM: {n_clients}",
    )
    ws0.cell(4, 1, "Карта объектов")
    for j, h in enumerate(["№", "Агент", "Название объекта", "Город", "Адрес", "Тип мойки"], 1):
        ws0.cell(5, j, h)
    ws0.cell(6, 1, 1)
    ws0.cell(6, 2, "Agent #365")
    ws0.cell(6, 3, obj_name)
    ws0.cell(6, 4, city)
    ws0.cell(6, 5, addr)
    ws0.cell(6, 6, "Самообслуживание")
    ws0.cell(13, 1, "Сводка по объектам и месяцам")
    wide_hdr = [
        "Название объекта",
        "Адрес",
        "Январь транзакций",
        "Январь визитов",
        "Январь клиентов",
        "Январь выручка",
        "Февраль транзакций",
        "Февраль визитов",
        "Февраль клиентов",
        "Февраль выручка",
        "Март транзакций",
        "Март визитов",
        "Март клиентов",
        "Март выручка",
        "Итого транзакций",
        "Итого визитов",
        "Итого клиентов",
        "Итого выручка",
        "Средний чек за визит",
    ]
    for j, h in enumerate(wide_hdr, 1):
        ws0.cell(14, j, h)
    ws0.cell(15, 1, obj_name)
    ws0.cell(15, 2, addr)
    ws0.cell(15, 3, mjan["tx"])
    ws0.cell(15, 4, mjan["vis"])
    ws0.cell(15, 5, jan_u)
    ws0.cell(15, 6, mjan["rev"])
    ws0.cell(15, 7, mfeb["tx"])
    ws0.cell(15, 8, mfeb["vis"])
    ws0.cell(15, 9, feb_u)
    ws0.cell(15, 10, mfeb["rev"])
    ws0.cell(15, 11, mmar["tx"])
    ws0.cell(15, 12, mmar["vis"])
    ws0.cell(15, 13, mar_u)
    ws0.cell(15, 14, mmar["rev"])
    ws0.cell(15, 15, total_tx)
    ws0.cell(15, 16, total_vis)
    ws0.cell(15, 17, total_cli)
    ws0.cell(15, 18, total_rev_all)
    ws0.cell(15, 19, round(avg_check_visit, 2))
    # ИТОГО ПО СЕТИ: при одном объекте совпадает со строкой объекта (как в файле Саранска)
    ws0.cell(21, 1, "ИТОГО ПО СЕТИ")
    ws0.cell(21, 2, None)
    ws0.cell(21, 3, mjan["tx"])
    ws0.cell(21, 4, mjan["vis"])
    ws0.cell(21, 5, jan_u)
    ws0.cell(21, 6, mjan["rev"])
    ws0.cell(21, 7, mfeb["tx"])
    ws0.cell(21, 8, mfeb["vis"])
    ws0.cell(21, 9, feb_u)
    ws0.cell(21, 10, mfeb["rev"])
    ws0.cell(21, 11, mmar["tx"])
    ws0.cell(21, 12, mmar["vis"])
    ws0.cell(21, 13, mar_u)
    ws0.cell(21, 14, mmar["rev"])
    ws0.cell(21, 15, total_tx)
    ws0.cell(21, 16, total_vis)
    ws0.cell(21, 17, total_cli)
    ws0.cell(21, 18, total_rev_all)
    ws0.cell(21, 19, round(avg_check_visit, 2))
    ws0.cell(23, 1, "Модель прогноза загрузки")
    ws0.cell(
        24,
        1,
        "RFM: R, F, M — квантили 1–5 по рангу среди клиентов Лейки; сегменты — Champions, Loyal, … (как в единой схеме Samara).",
    )

    # ----- RFM сегменты -----
    ws1 = wb.create_sheet("RFM сегменты")
    ws1.cell(1, 1, "RFM-сегментация клиентов")
    ws1.cell(
        3,
        1,
        f"RFM по всей выборке ({n_clients} клиентов). Сегменты — классическая 7-ячеечная карта (как AquaBot Samara).",
    )
    rfm_next = append_df(ws1, rfm_sheet_a, start_row=4, start_col=1)
    ws1.cell(rfm_next, 1, f"RFM: {obj_name} — {n_clients} клиентов")
    rfm2_next = append_df(ws1, rfm_sheet_b, start_row=rfm_next + 1, start_col=1)

    # ----- Клиенты для промо -----
    ws2 = wb.create_sheet("Клиенты для промо")
    ws2.cell(1, 1, "Список клиентов с промокодами и текстами сообщений")
    ws2.cell(
        2,
        1,
        "Сегмент — Champions / Loyal / …; промокоды по сегменту (см. лист «RFM сегменты»).",
    )
    append_df(ws2, promo_df, start_row=4, start_col=1)

    # ----- Погода и загрузка (заголовок и порядок колонок как в Саранске) -----
    ws3 = wb.create_sheet("Погода и загрузка")
    ws3.cell(1, 1, "Ежедневная загрузка и погода — Красноярск (Agent #365)")
    weather_out = pd.DataFrame(
        {
            "Дата": [pd.Timestamp(d).date() for d in load_w["date"]],
            "День недели": load_w["День недели"],
            "Визитов": load_w["визитов"],
            "Транзакций": load_w["Транзакций"].astype(int),
            "Транз/визит": np.round(load_w["Транз/визит"], 2),
            "Клиентов": load_w["клиентов"].astype(int),
            "Выручка": np.round(load_w["Выручка"], 2),
            "T мин °C": load_w["temp_min"],
            "T макс °C": load_w["temp_max"],
            "T средн °C": np.round(load_w["T средн °C"], 2),
            "Осадки мм": load_w["precipitation"],
            "Снегопад мм": load_w["snowfall"],
            "Выходной": load_w["Выходной"],
        }
    )
    append_df(ws3, weather_out, start_row=3, start_col=1)

    # ----- Когорты retention -----
    ws4 = wb.create_sheet("Когорты retention")
    ws4.cell(1, 1, "Когортный анализ — возвращаемость клиентов")
    ws4.cell(3, 1, "Вся сеть")
    ws4.cell(4, 1, "Показатель")
    ws4.cell(4, 2, "Значение")
    ws4.cell(4, 3, "Подробнее")
    cohort_rows = [
        ("Клиентов в январе", len(jan_phones), None),
        ("Клиентов в феврале", len(feb_phones), None),
        ("Клиентов в марте", len(mar_phones), None),
        ("Возврат январь→февраль", f"{ret_j_f:.1%}", f"{len(jan_phones & feb_phones)} из {len(jan_phones)}"),
        ("Возврат февраль→март", f"{ret_f_m:.1%}", f"{len(feb_phones & mar_phones)} из {len(feb_phones)}"),
        ("Ядро (все 3 месяца)", len(core), f"{len(core) / len(jan_phones):.1%} базы" if jan_phones else None),
        ("Новые в марте", len(mar_phones - (jan_phones | feb_phones)), None),
    ]
    for i, (a, b, c) in enumerate(cohort_rows, start=5):
        ws4.cell(i, 1, a)
        ws4.cell(i, 2, b)
        if c is not None:
            ws4.cell(i, 3, c)
    ws4.cell(13, 1, "По объектам")
    ws4.cell(14, 1, "Объект")
    ws4.cell(14, 2, "Январь")
    ws4.cell(14, 3, "Февраль")
    ws4.cell(14, 4, "Март")
    ws4.cell(14, 5, "Возврат Я→Ф")
    ws4.cell(14, 6, "Возврат Ф→М")
    ws4.cell(14, 7, "Ядро")
    ws4.cell(15, 1, obj_name)
    ws4.cell(15, 2, jan_u)
    ws4.cell(15, 3, feb_u)
    ws4.cell(15, 4, mar_u)
    ws4.cell(15, 5, ret_j_f)
    ws4.cell(15, 6, ret_f_m)
    ws4.cell(15, 7, len(core))

    # ----- Бюджет промо -----
    ws5 = wb.create_sheet("Бюджет промо")
    ws5.cell(1, 1, "Бюджет промо-кампаний с промокодами")
    ws5.cell(3, 1, "Стратегия промокодов")
    append_df(ws5, budget_df, start_row=4, start_col=1)
    br = 4 + len(budget_df) + 2
    ws5.cell(br, 1, "Пояснение механик промокодов")
    ws5.cell(br + 1, 1, "Механика")
    ws5.cell(br + 1, 2, "Как работает")
    ws5.cell(br + 1, 3, "Когда использовать")
    ws5.cell(
        br + 2,
        1,
        "Бонусы на счёт",
    )
    ws5.cell(br + 2, 2, "Клиент вводит промокод → бонусы зачисляются на счёт в приложении")
    ws5.cell(br + 2, 3, "Высокий квантиль R (недавние визиты)")
    ws5.cell(
        br + 3,
        1,
        "100% оплата промокодом",
    )
    ws5.cell(br + 3, 2, "Промокод покрывает мойку до указанной суммы")
    ws5.cell(br + 3, 3, "Низкий квантиль R — реактивация")
    ws5.cell(br + 5, 1, "Бюджет по объектам")
    ws5.cell(br + 6, 1, "Объект")
    ws5.cell(br + 6, 2, "Всего клиентов")
    ws5.cell(br + 6, 3, "Сегментов (классика RFM)")
    ws5.cell(br + 6, 4, "Бюджет (оценка)")
    ws5.cell(br + 7, 1, obj_name)
    ws5.cell(br + 7, 2, n_clients)
    ws5.cell(br + 7, 3, len(SEGMENT_ORDER))
    ws5.cell(br + 7, 4, total_bud)

    ref_xlsx = root / "H2O_Saransk_Promo_Plan.xlsx"
    apply_visual_style_saransk(
        wb,
        ref_xlsx,
        rfm_next=rfm_next,
        rfm2_next=rfm2_next,
    )

    wb.save(out_xlsx)

    lines = [
        "# Agent #365 (ИП Чепкасова Н. А.) — предиктивка и RFM",
        "",
        f"**Период:** {d0} — {d1}",
        "",
        "## Метод RFM (реальный)",
        "- Баллы **R, F, M** от 1 до 5 — **квантили по рангу** среди клиентов Лейки этого агента за период.",
        "- **R** (recency): чем меньше дней с последнего визита, тем выше балл.",
        "- **F** (frequency): чем больше визитов, тем выше балл.",
        "- **M** (monetary): чем выше сумма (LTV по визитам), тем выше балл.",
        "- **Сегменты** в Excel — **Champions, Loyal, Promising, …** (единая схема с `AquaBot_Samara_PREDECTIVE.xlsx` / промо-планом сети).",
        "",
        "## Объект",
        f"- {obj_name}, {addr}",
        "",
        "## Ключевые цифры (Лейка)",
        f"- Клиентов: **{n_clients}**",
        f"- Визитов: **{len(visits)}**",
        f"- Выручка: **{total_rev_all:,.0f} ₽**".replace(",", " "),
        "",
        "## Файлы",
        f"- `{out_xlsx.name}` — структура как у `H2O_Saransk_Promo_Plan.xlsx`",
        f"- `PROMO_PLAN_SPEC.md` — универсальная методика промо-плана и RFM (не привязана к номеру агента или числу моек)",
        "",
        f"*Сгенерировано: {datetime.now().strftime('%Y-%m-%d %H:%M')}*",
    ]
    out_md.write_text("\n".join(lines), encoding="utf-8")

    print(f"OK: {out_xlsx}")
    print(f"OK: {out_md}")


if __name__ == "__main__":
    main()
