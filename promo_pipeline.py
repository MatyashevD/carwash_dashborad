"""
Универсальный пайплайн: CSV Лейки → RFM → сегменты → промо-план Excel (BytesIO).

Не привязан к конкретному агенту/городу — всё задаётся через PromoConfig.
Методика: PROMO_PLAN_SPEC.md.
"""
from __future__ import annotations

import io
import json
import re
import urllib.request
from dataclasses import dataclass, field
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------------------------------------------------------------------
# Константы
# ---------------------------------------------------------------------------

YANDEX_PHONE = "133133133133"
TBANK_PHONE = "71119999991"
TIME_WINDOW_MIN = 30

WD_RU = (
    "Понедельник", "Вторник", "Среда", "Четверг",
    "Пятница", "Суббота", "Воскресенье",
)

SEGMENT_ORDER: tuple[str, ...] = (
    "Champions", "Loyal", "Promising", "Need Attention",
    "At Risk", "About to Sleep", "Hibernating",
)

INITIATIVE_RU: dict[str, str] = {
    "Champions": "Удержание лучших",
    "Loyal": "Рост частоты",
    "Promising": "Конверсия новых",
    "Need Attention": "Лёгкая реактивация",
    "At Risk": "Срочная реактивация",
    "About to Sleep": "Реактивация уходящих",
    "Hibernating": "Возврат потерянных",
}

_DEFAULT_R_BY_SEGMENT: dict[str, int] = {
    "Champions": 5, "Loyal": 4, "Promising": 5, "Need Attention": 3,
    "At Risk": 2, "About to Sleep": 2, "Hibernating": 1,
}

_PROMO_TEMPLATES: dict[str, tuple[str, str, str, str]] = {
    "Champions": ("VIP50{sfx}",   "Бонусы на счёт",       "50₽ бонус на счёт",           "1 — Высший (удержание VIP)"),
    "Loyal":     ("PLUS30{sfx}",  "Бонусы на счёт",       "30₽ бонус на счёт",           "2 — Высокий (рост частоты)"),
    "Promising": ("START50{sfx}", "Бонусы на счёт",       "50₽ бонус на счёт",           "3 — Средний (конверсия)"),
    "Need Attention": ("MISS100{sfx}",  "100% оплата промокодом", "Бесплатная мойка до 100₽", "4 — Высокий (реактивация)"),
    "At Risk":        ("BACK150{sfx}",  "100% оплата промокодом", "Бесплатная мойка до 150₽", "5 — Очень высокий (срочная реактивация)"),
    "About to Sleep": ("WAKE100{sfx}",  "100% оплата промокодом", "Бесплатная мойка до 100₽", "6 — Средний (реактивация)"),
    "Hibernating":    ("SPRING200{sfx}","100% оплата промокодом", "Бесплатная мойка до 200₽", "7 — Низкий (холодная реактивация)"),
}


# ---------------------------------------------------------------------------
# Конфигурация партнёра
# ---------------------------------------------------------------------------

@dataclass
class PromoConfig:
    partner_display_name: str
    partner_filter: str            # regex для поля «Партнёр»
    city: str = ""
    address_short: str = ""        # «на Светлогорской» — для push
    promo_suffix: str = ""         # KR365 и т.п.
    weather_coords: tuple[float, float] | None = None
    include_weather: bool = True
    wash_type: str = "Мойка автомобиля"
    # Шаблон эталона: при наличии подтягиваем ширины и стили
    ref_wb_bytes: bytes | None = field(default=None, repr=False)


# ---------------------------------------------------------------------------
# Утилиты для данных
# ---------------------------------------------------------------------------

def load_order_table_from_bytes(buf: io.BytesIO) -> pd.DataFrame:
    buf.seek(0)
    df = pd.read_csv(buf, sep=";", encoding="utf-8-sig", dtype=str)
    for col in ("Поступило на бокс", "Оплачено деньгами", "Оплачено бонусами", "Начислено кешбека"):
        if col in df.columns:
            s = (
                df[col].astype(str)
                .str.replace("\u00a0", " ", regex=False)
                .str.replace(" ", "", regex=False)
                .str.replace(",", ".", regex=False)
            )
            df[col] = pd.to_numeric(s, errors="coerce").fillna(0.0)
    df["Дата оплаты"] = pd.to_datetime(df["Дата оплаты"], errors="coerce")
    df["date"] = df["Дата оплаты"].dt.date
    df["total"] = df.get("Оплачено деньгами", 0) + df.get("Оплачено бонусами", 0)
    return df


def normalize_phone(phone: str) -> str:
    if pd.isna(phone) or str(phone).strip() == "":
        return ""
    s = str(phone).strip().replace(",", ".")
    try:
        return str(int(float(s)))
    except (ValueError, OverflowError):
        return "".join(filter(str.isdigit, s))


def categorize_phone(phone: str) -> str:
    p = normalize_phone(phone)
    if p == YANDEX_PHONE:
        return "Яндекс"
    if p == TBANK_PHONE:
        return "Т-Банк"
    return "Лейка"


def group_transactions_to_visits(df: pd.DataFrame) -> pd.DataFrame:
    empty = pd.DataFrame(columns=[
        "visit_id", "visit_total", "Оплачено деньгами",
        "Оплачено бонусами", "Поступило на бокс", "Телефон", "Дата оплаты",
    ])
    if df.empty or "Дата оплаты" not in df.columns or "Телефон" not in df.columns:
        return empty
    w = df.dropna(subset=["Дата оплаты", "Телефон"]).copy()
    if w.empty:
        return empty
    w = w.sort_values(["Телефон", "Дата оплаты"])
    prev = w.groupby("Телефон")["Дата оплаты"].shift(1)
    gap = (w["Дата оплаты"] - prev).dt.total_seconds() / 60
    w["visit_id"] = (gap.isna() | (gap > TIME_WINDOW_MIN)).cumsum()
    visits = (
        w.groupby("visit_id")
        .agg(total=("total", "sum"), **{
            "Оплачено деньгами": ("Оплачено деньгами", "sum"),
            "Оплачено бонусами": ("Оплачено бонусами", "sum"),
            "Поступило на бокс": ("Поступило на бокс", "sum"),
            "Телефон": ("Телефон", "first"),
            "Дата оплаты": ("Дата оплаты", "first"),
        })
        .reset_index()
        .rename(columns={"total": "visit_total"})
    )
    return visits


# ---------------------------------------------------------------------------
# RFM
# ---------------------------------------------------------------------------

def quintile_score_1_to_5(s: pd.Series, *, higher_is_better: bool) -> pd.Series:
    rank = s.rank(method="first")
    if not higher_is_better:
        rank = rank.max() + 1 - rank
    n = len(s)
    if n == 0:
        return pd.Series(dtype=int)
    pct = (rank - 0.5) / n
    scores = np.clip(np.ceil(pct * 5).astype(int), 1, 5)
    return pd.Series(scores.values, index=s.index, dtype=int)


def marketing_segment(r: int, f: int, m: int) -> str:
    if r >= 4 and f >= 4 and m >= 4:
        return "Champions"
    if r >= 4 and f >= 4:
        return "Loyal"
    if r >= 4 and f <= 2:
        return "Promising"
    if r >= 4:
        return "Loyal"
    if r == 3 and f >= 3 and m >= 3:
        return "Need Attention"
    if r == 2 and f >= 3 and m >= 3:
        return "At Risk"
    if r == 2:
        return "About to Sleep"
    if r == 3:
        return "About to Sleep"
    return "Hibernating"


def _promo_bundle(segment: str, suffix: str) -> tuple[str, str, str, str]:
    tpl = _PROMO_TEMPLATES[segment]
    code = tpl[0].format(sfx=suffix)
    return code, tpl[1], tpl[2], tpl[3]


def _rep_r(clients: pd.DataFrame, segment: str) -> int:
    sub = clients[clients["Сегмент"] == segment]["R"]
    return int(round(float(sub.median()))) if len(sub) else _DEFAULT_R_BY_SEGMENT[segment]


def _budget_estimate(r: int, n: int) -> tuple[int, str, str]:
    if r >= 4:
        return n * 50, "Ежемесячно", "~оценка удержания"
    if r == 3:
        return n * 35, "Ежемесячно", "~оценка частоты"
    if r == 2:
        return n * 100, "Однократно", "~оценка реактивации"
    return n * 200, "Однократно", "~оценка возврата"


# ---------------------------------------------------------------------------
# Push-тексты (пачки, без имени)
# ---------------------------------------------------------------------------

def _push_text(segment: str, pcode: str, summa: str, mech: str, rec: str, location: str) -> str:
    loc = f" {location}" if location else ""
    if mech == "Бонусы на счёт":
        amt = "VIP-бонус 50₽" if segment == "Champions" else ("бонус 30₽" if segment == "Loyal" else "бонус 50₽")
        tail = "Спасибо за вашу лояльность." if segment != "Promising" else "Будем рады снова увидеть вас на мойке!"
        return f"Спасибо за выбор мойки{loc}! Вам начислен {amt} на бонусный счёт. {tail}"
    tail = summa[:1].lower() + summa[1:] if summa else ""
    return f"Спасибо за выбор мойки{loc}! Активируйте промокод {pcode} в приложении — {tail}. ({rec})"


# ---------------------------------------------------------------------------
# Погода (Open-Meteo Archive, timezone=auto)
# ---------------------------------------------------------------------------

def _fetch_weather(lat: float, lon: float, d0, d1) -> pd.DataFrame | None:
    url = (
        f"https://archive-api.open-meteo.com/v1/archive?"
        f"latitude={lat}&longitude={lon}"
        f"&start_date={d0.isoformat()}&end_date={d1.isoformat()}"
        f"&daily=temperature_2m_max,temperature_2m_min,precipitation_sum,"
        f"snowfall_sum,rain_sum,weathercode&timezone=auto"
    )
    try:
        with urllib.request.urlopen(url, timeout=60) as resp:
            data = json.loads(resp.read().decode())
        daily = data["daily"]
        return pd.DataFrame({
            "date": pd.to_datetime(pd.Series(daily["time"])).dt.date,
            "temp_max": daily["temperature_2m_max"],
            "temp_min": daily["temperature_2m_min"],
            "precipitation": daily["precipitation_sum"],
            "snowfall": daily["snowfall_sum"],
            "rain": daily["rain_sum"],
            "weathercode": daily["weathercode"],
        })
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Excel-утилиты
# ---------------------------------------------------------------------------

def _append_df(ws, df: pd.DataFrame, start_row: int, start_col: int = 1) -> int:
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        for c_idx, val in enumerate(row):
            ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=val)
    return start_row + len(df) + 1


def _apply_visual_style(wb: Workbook, *, rfm_next: int, rfm2_next: int,
                         ref_wb_bytes: bytes | None, has_weather: bool) -> None:
    thin = Side(style="thin", color="FFD0D0D0")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="FF4472C4")
    hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=11, bold=True)
    fill_bonus = PatternFill("solid", fgColor="FFE2EFDA")
    fill_promo = PatternFill("solid", fgColor="FFFCE4D6")

    if ref_wb_bytes:
        from openpyxl import load_workbook as _lwb
        ref = _lwb(io.BytesIO(ref_wb_bytes))
        try:
            for name in wb.sheetnames:
                if name not in ref.sheetnames:
                    continue
                tw, rw = wb[name], ref[name]
                for col, dim in rw.column_dimensions.items():
                    if dim.width:
                        tw.column_dimensions[col].width = dim.width
        finally:
            ref.close()
    else:
        fallback = {
            "Обзор и модель": {"A": 55, "B": 36, "S": 23},
            "RFM сегменты": {"A": 46, "K": 55},
            "Клиенты для промо": {"A": 55, "N": 55},
            "Погода и загрузка": {"A": 42, "M": 40},
            "Когорты retention": {"A": 45},
            "Бюджет промо": {"B": 55, "K": 55},
        }
        for sn, cols in fallback.items():
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            for letter, w in cols.items():
                ws.column_dimensions[letter].width = w

    def paint_header(ws, row, max_col):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.border = grid
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def paint_grid(ws, r1, r2, c1, c2):
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                ws.cell(row=r, column=c).border = grid

    ws0 = wb["Обзор и модель"]
    paint_header(ws0, 5, 6)
    paint_header(ws0, 14, 19)
    paint_grid(ws0, 6, 6, 1, 6)
    paint_grid(ws0, 15, 15, 1, 19)
    paint_grid(ws0, 21, 21, 1, 19)
    ws0.cell(4, 1).font = title_font
    ws0.cell(13, 1).font = title_font
    for r in (1, 2):
        ws0.cell(r, 1).alignment = Alignment(wrap_text=True)

    ws1 = wb["RFM сегменты"]
    paint_header(ws1, 4, 11)
    paint_grid(ws1, 5, rfm_next - 1, 1, 11)
    ws1.cell(rfm_next, 1).font = title_font
    hr2 = rfm_next + 1
    paint_header(ws1, hr2, 6)
    paint_grid(ws1, hr2 + 1, rfm2_next - 1, 1, 6)
    ws1.freeze_panes = "A5"

    ws2 = wb["Клиенты для промо"]
    paint_header(ws2, 4, 15)
    for r in range(5, ws2.max_row + 1):
        mech = ws2.cell(r, 11).value
        m_str = str(mech or "")
        row_fill = None
        if "Бонусы" in m_str:
            row_fill = fill_bonus
        elif "100%" in m_str or "промокод" in m_str.lower():
            row_fill = fill_promo
        for c in range(1, 16):
            cell = ws2.cell(r, c)
            cell.border = grid
            if row_fill is not None and c in (11, 12):
                cell.fill = row_fill
    ws2.freeze_panes = "A5"

    if has_weather and "Погода и загрузка" in wb.sheetnames:
        ws3 = wb["Погода и загрузка"]
        paint_header(ws3, 3, 13)
        paint_grid(ws3, 4, ws3.max_row, 1, 13)
        ws3.freeze_panes = "A4"

    ws4 = wb["Когорты retention"]
    paint_header(ws4, 4, 3)
    paint_grid(ws4, 5, 11, 1, 3)
    ws4.cell(13, 1).font = title_font
    paint_header(ws4, 14, 7)
    paint_grid(ws4, 15, 15, 1, 7)

    ws5 = wb["Бюджет промо"]
    paint_header(ws5, 4, 11)
    for r in range(5, ws5.max_row + 1):
        for c in range(1, 12):
            if ws5.cell(r, c).value is not None or c <= 2:
                ws5.cell(r, c).border = grid


# ---------------------------------------------------------------------------
# Генерация промо-плана
# ---------------------------------------------------------------------------

def generate_promo_plan(
    config: PromoConfig,
    csv_buffers: list[io.BytesIO],
    *,
    progress_cb=None,
) -> io.BytesIO:
    """
    Основной вход. Возвращает BytesIO с готовым .xlsx.
    progress_cb(float) — опциональный колбэк 0.0..1.0 для progress bar.
    """
    def _progress(v: float):
        if progress_cb:
            progress_cb(v)

    _progress(0.05)

    # 1. Загрузка и объединение
    raw = pd.concat([load_order_table_from_bytes(b) for b in csv_buffers], ignore_index=True)
    raw["partner_category"] = raw["Телефон"].apply(categorize_phone)
    _progress(0.15)

    # 2. Фильтрация по партнёру
    pat = re.compile(config.partner_filter, re.I)
    agent_mask = raw["Партнёр"].astype(str).apply(lambda s: bool(pat.search(s)))
    agent_df = raw[agent_mask].copy()
    wash = agent_df[agent_df["Тип оплаты"].astype(str).str.strip() == config.wash_type].copy()
    leyka = wash[wash["partner_category"] == "Лейка"].copy()

    if leyka.empty:
        raise ValueError(f"Нет данных Лейки по партнёру «{config.partner_display_name}» (фильтр: {config.partner_filter})")

    analysis_end = pd.Timestamp(leyka["Дата оплаты"].max()).normalize()
    _progress(0.20)

    # 3. Визиты и клиентская база
    visits = group_transactions_to_visits(leyka)
    last_visit = visits.groupby("Телефон")["Дата оплаты"].max()
    vc = visits.groupby("Телефон").size()
    ltv = visits.groupby("Телефон")["visit_total"].sum()
    clients = pd.DataFrame({
        "Телефон": ltv.index,
        "LTV": ltv.values,
        "визитов": vc.reindex(ltv.index).values,
        "последний_визит": last_visit.reindex(ltv.index).values,
    }).reset_index(drop=True)
    clients["recency_дней"] = (analysis_end - pd.to_datetime(clients["последний_визит"])).dt.days.astype(int)
    _progress(0.30)

    # 4. RFM
    clients["R"] = quintile_score_1_to_5(clients["recency_дней"], higher_is_better=False)
    clients["F"] = quintile_score_1_to_5(clients["визитов"].astype(float), higher_is_better=True)
    clients["M"] = quintile_score_1_to_5(clients["LTV"].astype(float), higher_is_better=True)
    clients["Сегмент"] = [
        marketing_segment(int(r), int(f), int(m))
        for r, f, m in zip(clients["R"], clients["F"], clients["M"])
    ]

    last_tx = (
        leyka.sort_values("Дата оплаты")
        .groupby("Телефон").tail(1)[["Телефон", "Клиент"]]
        .rename(columns={"Клиент": "Имя клиента"})
    )
    clients = clients.merge(last_tx, on="Телефон", how="left")
    total_rev = clients["LTV"].sum()
    n_clients = len(clients)
    _progress(0.40)

    sfx = config.promo_suffix

    # 5. RFM-агрегат по 7 сегментам
    seg_counts = clients["Сегмент"].value_counts()
    seg_rev = clients.groupby("Сегмент", observed=False)["LTV"].sum()
    rfm_rows: list[dict] = []
    for seg in SEGMENT_ORDER:
        n = int(seg_counts.get(seg, 0))
        rev = float(seg_rev.get(seg, 0.0)) if n else 0.0
        sub = clients[clients["Сегмент"] == seg]
        pcode, mech, sum_txt, rec = _promo_bundle(seg, sfx)
        rfm_rows.append({
            "Сегмент": seg, "Клиентов": n,
            "Доля базы": n / n_clients if n_clients else 0.0,
            "Выручка": rev,
            "Доля выручки": rev / total_rev if total_rev else 0.0,
            "Ср. визитов": float(sub["визитов"].mean()) if n else 0.0,
            "Ср. дней с посл. визита": float(sub["recency_дней"].mean()) if n else 0.0,
            "Промокод": pcode, "Механика промокода": mech,
            "Сумма": sum_txt, "Рекомендуемое действие": rec,
        })
    rfm_agg = pd.DataFrame(rfm_rows)
    _progress(0.45)

    # 6. Когорты (автоопределение месяцев из данных)
    leyka["month"] = pd.to_datetime(leyka["date"]).dt.to_period("M").astype(str)
    months_sorted = sorted(leyka["month"].unique())

    phone_sets: dict[str, set] = {}
    for m in months_sorted:
        phone_sets[m] = set(leyka[leyka["month"] == m]["Телефон"].unique())

    visits["month"] = visits["Дата оплаты"].dt.to_period("M").astype(str)

    obj_names = wash["Автомойка"].unique().tolist() if "Автомойка" in wash.columns else []
    obj_name = obj_names[0] if obj_names else ""
    addr = wash["Адрес"].iloc[0] if len(wash) and "Адрес" in wash.columns else ""
    city = config.city or (str(addr).split(",")[0].strip() if addr else "")

    def month_stats(m_str: str):
        lv = leyka[leyka["month"] == m_str]
        vv = visits[visits["month"] == m_str]
        return {"tx": len(lv), "vis": len(vv), "cli": lv["Телефон"].nunique(), "rev": lv["total"].sum()}

    m_stats = {m: month_stats(m) for m in months_sorted}
    total_tx = sum(s["tx"] for s in m_stats.values())
    total_vis = sum(s["vis"] for s in m_stats.values())
    total_rev_all = float(leyka["total"].sum())
    avg_check_visit = total_rev_all / len(visits) if len(visits) else 0
    _progress(0.50)

    # 7. Погода (опционально)
    d0, d1 = wash["date"].min(), wash["date"].max()
    weather_df = None
    has_weather = False
    if config.include_weather and config.weather_coords:
        lat, lon = config.weather_coords
        weather_df = _fetch_weather(lat, lon, d0, d1)
    _progress(0.60)

    load_w = None
    if weather_df is not None:
        has_weather = True
        daily_vis = visits.copy()
        daily_vis["d"] = daily_vis["Дата оплаты"].dt.date
        daily_v = daily_vis.groupby("d").agg(визитов=("visit_id", "count"), клиентов=("Телефон", "nunique"))
        daily_tx = leyka.groupby(leyka["Дата оплаты"].dt.date).agg(Выручка=("total", "sum"), Транзакций=("total", "size"))
        load_w = weather_df.merge(daily_v, left_on="date", right_index=True, how="left")
        load_w = load_w.merge(daily_tx, left_on="date", right_index=True, how="left")
        load_w["визитов"] = load_w["визитов"].fillna(0).astype(int)
        load_w["клиентов"] = load_w["клиентов"].fillna(0).astype(int)
        load_w["Транзакций"] = load_w["Транзакций"].fillna(0).astype(int)
        load_w["Выручка"] = load_w["Выручка"].fillna(0.0)
        load_w["Транз/визит"] = np.where(load_w["визитов"] > 0, load_w["Транзакций"] / load_w["визитов"], 0.0)
        load_w["T средн °C"] = (load_w["temp_max"] + load_w["temp_min"]) / 2
        _dt = pd.to_datetime(load_w["date"])
        load_w["День недели"] = [WD_RU[_dt.iloc[i].dayofweek] for i in range(len(load_w))]
        load_w["Выходной"] = _dt.dt.weekday.ge(5).map({True: "Да", False: "Нет"})
    _progress(0.65)

    # 8. Клиенты для промо
    loc = config.address_short
    promo_rows = []
    for _, row in clients.iterrows():
        seg = str(row["Сегмент"])
        pcode, mech, summa, rec = _promo_bundle(seg, sfx)
        name = str(row.get("Имя клиента") or "").strip()
        objs = ", ".join(obj_names) if len(obj_names) > 1 else (f"{obj_name}, {addr}" if addr else obj_name)
        promo_rows.append({
            "Телефон": row["Телефон"],
            "Имя клиента": name,
            "Объекты посещения": objs,
            "Сегмент": seg,
            "LTV (общая сумма оплат)": row["LTV"],
            "Количество визитов": int(row["визитов"]),
            "Количество транзакций": int(leyka[leyka["Телефон"] == row["Телефон"]].shape[0]),
            "Дата последнего визита": pd.to_datetime(row["последний_визит"]).strftime("%d.%m.%Y"),
            "Дней с последнего визита": int(row["recency_дней"]),
            "Промокод": pcode,
            "Механика": mech,
            "Сумма промо": summa,
            "Канал": "Push",
            "Текст сообщения": _push_text(seg, pcode, summa, mech, rec, loc),
            "Приоритет": f"{seg} — {INITIATIVE_RU.get(seg, seg)}",
        })
    seg_rank = {s: i for i, s in enumerate(SEGMENT_ORDER)}
    promo_df = pd.DataFrame(promo_rows)
    promo_df["_sr"] = promo_df["Сегмент"].map(seg_rank)
    promo_df = promo_df.sort_values(["_sr", "LTV (общая сумма оплат)"], ascending=[True, False]).drop(columns=["_sr"])
    _progress(0.70)

    # 9. Бюджет
    budget_rows = []
    for seg in SEGMENT_ORDER:
        gr = rfm_agg[rfm_agg["Сегмент"] == seg].iloc[0]
        n = int(gr["Клиентов"])
        r_rep = _rep_r(clients, seg)
        pcode, mech, sum_txt, _ = _promo_bundle(seg, sfx)
        bud, period, roi_hint = _budget_estimate(r_rep, n) if n else (0, "—", "—")
        budget_rows.append({
            "Инициатива": INITIATIVE_RU[seg], "Сегмент": seg, "Промокод": pcode,
            "Механика": mech, "Сумма": sum_txt, "Клиентов": n,
            "Бюджет": int(bud), "Периодичность": period,
            "Ожидаемый доп. доход/мес": roi_hint, "Окупаемость": "—",
            "Текст сообщения": (
                f"Начисление бонусов (код кампании {pcode}) — см. «Клиенты для промо»"
                if mech == "Бонусы на счёт"
                else f"Промокод на мойку {pcode} — см. «Клиенты для промо»"
            ),
        })
    total_bud = sum(b["Бюджет"] for b in budget_rows)
    budget_rows.append({
        "Инициатива": "ИТОГО", "Сегмент": None, "Промокод": None,
        "Механика": None, "Сумма": None, "Клиентов": n_clients,
        "Бюджет": total_bud, "Периодичность": None,
        "Ожидаемый доп. доход/мес": "~оценка", "Окупаемость": "< 2–3 мес (оценка)",
        "Текст сообщения": None,
    })
    budget_df = pd.DataFrame(budget_rows)
    _progress(0.75)

    # 10. Сборка Excel
    days_count = (pd.Timestamp(d1) - pd.Timestamp(d0)).days + 1
    unique_months_cli = {m: len(ps) for m, ps in phone_sets.items()}

    rfm_sheet_a = rfm_agg[[
        "Сегмент", "Клиентов", "Доля базы", "Выручка", "Доля выручки",
        "Ср. визитов", "Ср. дней с посл. визита",
        "Промокод", "Механика промокода", "Сумма", "Рекомендуемое действие",
    ]]
    rfm_sheet_b = rfm_agg[["Сегмент", "Клиентов", "Доля базы", "Выручка", "Ср. визитов", "Ср. дней с посл. визита"]].rename(
        columns={"Доля базы": "Доля"}
    )

    wb = Workbook()

    # --- Обзор и модель ---
    ws0 = wb.active
    ws0.title = "Обзор и модель"
    n_objs = len(obj_names)
    ws0.cell(1, 1, f"Предиктивная аналитика — {config.partner_display_name}, {city}")
    ws0.cell(2, 1, f"Данные: {d0} — {d1} ({days_count} дней), {n_objs or 1} объект(ов), клиентов Лейки: {n_clients}")
    ws0.cell(4, 1, "Карта объектов")
    for j, h in enumerate(["№", "Агент", "Название объекта", "Город", "Адрес", "Тип мойки"], 1):
        ws0.cell(5, j, h)
    for idx, on in enumerate(obj_names or [obj_name], start=1):
        ws0.cell(5 + idx, 1, idx)
        ws0.cell(5 + idx, 2, config.partner_display_name)
        ws0.cell(5 + idx, 3, on)
        ws0.cell(5 + idx, 4, city)
        ws0.cell(5 + idx, 5, addr)
        ws0.cell(5 + idx, 6, "Самообслуживание")

    ws0.cell(13, 1, "Сводка по объектам и месяцам")

    month_labels = []
    for m_str in months_sorted:
        try:
            dt = pd.Timestamp(m_str)
            month_labels.append(dt.strftime("%B").capitalize())
        except Exception:
            month_labels.append(m_str)

    wide_hdr = ["Название объекта", "Адрес"]
    for ml in month_labels:
        wide_hdr.extend([f"{ml} транзакций", f"{ml} визитов", f"{ml} клиентов", f"{ml} выручка"])
    wide_hdr.extend(["Итого транзакций", "Итого визитов", "Итого клиентов", "Итого выручка", "Средний чек за визит"])
    for j, h in enumerate(wide_hdr, 1):
        ws0.cell(14, j, h)

    ws0.cell(15, 1, obj_name)
    ws0.cell(15, 2, addr)
    col = 3
    for m_str in months_sorted:
        ms = m_stats[m_str]
        ws0.cell(15, col, ms["tx"]); col += 1
        ws0.cell(15, col, ms["vis"]); col += 1
        ws0.cell(15, col, unique_months_cli.get(m_str, 0)); col += 1
        ws0.cell(15, col, ms["rev"]); col += 1
    ws0.cell(15, col, total_tx); col += 1
    ws0.cell(15, col, total_vis); col += 1
    ws0.cell(15, col, n_clients); col += 1
    ws0.cell(15, col, total_rev_all); col += 1
    ws0.cell(15, col, round(avg_check_visit, 2))

    ws0.cell(21, 1, "ИТОГО ПО СЕТИ")
    col = 3
    for m_str in months_sorted:
        ms = m_stats[m_str]
        ws0.cell(21, col, ms["tx"]); col += 1
        ws0.cell(21, col, ms["vis"]); col += 1
        ws0.cell(21, col, unique_months_cli.get(m_str, 0)); col += 1
        ws0.cell(21, col, ms["rev"]); col += 1
    ws0.cell(21, col, total_tx); col += 1
    ws0.cell(21, col, total_vis); col += 1
    ws0.cell(21, col, n_clients); col += 1
    ws0.cell(21, col, total_rev_all); col += 1
    ws0.cell(21, col, round(avg_check_visit, 2))

    ws0.cell(23, 1, "Модель прогноза загрузки")
    ws0.cell(24, 1, "RFM: R,F,M — квантили 1–5; сегменты — Champions, Loyal, … (единая схема).")
    _progress(0.80)

    # --- RFM сегменты ---
    ws1 = wb.create_sheet("RFM сегменты")
    ws1.cell(1, 1, "RFM-сегментация клиентов")
    ws1.cell(3, 1, f"RFM по всей выборке ({n_clients} клиентов). Сегменты — классическая 7-ячеечная карта.")
    rfm_next = _append_df(ws1, rfm_sheet_a, start_row=4)
    ws1.cell(rfm_next, 1, f"RFM: {obj_name} — {n_clients} клиентов")
    rfm2_next = _append_df(ws1, rfm_sheet_b, start_row=rfm_next + 1)

    # --- Клиенты для промо ---
    ws2 = wb.create_sheet("Клиенты для промо")
    ws2.cell(1, 1, "Список клиентов с промокодами и текстами сообщений")
    ws2.cell(2, 1, "Сегмент — Champions / Loyal / …; промокоды по сегменту (см. лист «RFM сегменты»).")
    _append_df(ws2, promo_df, start_row=4)
    _progress(0.85)

    # --- Погода и загрузка (опционально) ---
    if has_weather and load_w is not None:
        ws3 = wb.create_sheet("Погода и загрузка")
        ws3.cell(1, 1, f"Ежедневная загрузка и погода — {city} ({config.partner_display_name})")
        weather_out = pd.DataFrame({
            "Дата": [pd.Timestamp(d).date() for d in load_w["date"]],
            "День недели": load_w["День недели"],
            "Визитов": load_w["визитов"],
            "Транзакций": load_w["Транзакций"].astype(int),
            "Транз/визит": np.round(load_w["Транз/визит"], 2),
            "Клиентов": load_w["клиентов"].astype(int),
            "Выручка": np.round(load_w["Выручка"], 2),
            "T мин °C": load_w["temp_min"],
            "T макс °C": load_w["temp_max"],
            "T средн °C": np.round(load_w["T средн °C"], 2),
            "Осадки мм": load_w["precipitation"],
            "Снегопад мм": load_w["snowfall"],
            "Выходной": load_w["Выходной"],
        })
        _append_df(ws3, weather_out, start_row=3)

    # --- Когорты retention ---
    ws4 = wb.create_sheet("Когорты retention")
    ws4.cell(1, 1, "Когортный анализ — возвращаемость клиентов")
    ws4.cell(3, 1, "Вся сеть")
    ws4.cell(4, 1, "Показатель"); ws4.cell(4, 2, "Значение"); ws4.cell(4, 3, "Подробнее")

    row_i = 5
    for m_str in months_sorted:
        ws4.cell(row_i, 1, f"Клиентов ({m_str})")
        ws4.cell(row_i, 2, len(phone_sets[m_str]))
        row_i += 1

    pairs = list(zip(months_sorted, months_sorted[1:]))
    for m_a, m_b in pairs:
        ps_a, ps_b = phone_sets[m_a], phone_sets[m_b]
        ret = len(ps_a & ps_b) / len(ps_a) if ps_a else 0
        ws4.cell(row_i, 1, f"Возврат {m_a}→{m_b}")
        ws4.cell(row_i, 2, f"{ret:.1%}")
        ws4.cell(row_i, 3, f"{len(ps_a & ps_b)} из {len(ps_a)}")
        row_i += 1

    if len(months_sorted) >= 3:
        core = set.intersection(*phone_sets.values())
        first_set = phone_sets[months_sorted[0]]
        ws4.cell(row_i, 1, f"Ядро (все {len(months_sorted)} мес.)")
        ws4.cell(row_i, 2, len(core))
        ws4.cell(row_i, 3, f"{len(core) / len(first_set):.1%} базы" if first_set else None)
        row_i += 1

    last_month_set = phone_sets[months_sorted[-1]] if months_sorted else set()
    earlier = set.union(*(phone_sets[m] for m in months_sorted[:-1])) if len(months_sorted) > 1 else set()
    ws4.cell(row_i, 1, f"Новые в {months_sorted[-1]}" if months_sorted else "Новые")
    ws4.cell(row_i, 2, len(last_month_set - earlier))

    ws4.cell(13, 1, "По объектам")
    ws4.cell(14, 1, "Объект")
    for j, m_str in enumerate(months_sorted, 2):
        ws4.cell(14, j, m_str)
    ret_col = 2 + len(months_sorted)
    for j, (m_a, m_b) in enumerate(pairs):
        ws4.cell(14, ret_col + j, f"Возврат {m_a[:4]}→{m_b[-2:]}")
    core_col = ret_col + len(pairs)
    ws4.cell(14, core_col, "Ядро")

    ws4.cell(15, 1, obj_name)
    for j, m_str in enumerate(months_sorted, 2):
        ws4.cell(15, j, unique_months_cli.get(m_str, 0))
    for j, (m_a, m_b) in enumerate(pairs):
        ps_a, ps_b = phone_sets[m_a], phone_sets[m_b]
        ws4.cell(15, ret_col + j, len(ps_a & ps_b) / len(ps_a) if ps_a else 0)
    if len(months_sorted) >= 3:
        ws4.cell(15, core_col, len(core))
    _progress(0.90)

    # --- Бюджет промо ---
    ws5 = wb.create_sheet("Бюджет промо")
    ws5.cell(1, 1, "Бюджет промо-кампаний с промокодами")
    ws5.cell(3, 1, "Стратегия промокодов")
    _append_df(ws5, budget_df, start_row=4)
    br = 4 + len(budget_df) + 2
    ws5.cell(br, 1, "Пояснение механик промокодов")
    ws5.cell(br + 1, 1, "Механика"); ws5.cell(br + 1, 2, "Как работает"); ws5.cell(br + 1, 3, "Когда использовать")
    ws5.cell(br + 2, 1, "Бонусы на счёт")
    ws5.cell(br + 2, 2, "Клиент вводит промокод → бонусы зачисляются на счёт в приложении")
    ws5.cell(br + 2, 3, "Высокий квантиль R (недавние визиты)")
    ws5.cell(br + 3, 1, "100% оплата промокодом")
    ws5.cell(br + 3, 2, "Промокод покрывает мойку до указанной суммы")
    ws5.cell(br + 3, 3, "Низкий квантиль R — реактивация")
    ws5.cell(br + 5, 1, "Бюджет по объектам")
    ws5.cell(br + 6, 1, "Объект"); ws5.cell(br + 6, 2, "Всего клиентов")
    ws5.cell(br + 6, 3, "Сегментов (классика RFM)"); ws5.cell(br + 6, 4, "Бюджет (оценка)")
    ws5.cell(br + 7, 1, obj_name); ws5.cell(br + 7, 2, n_clients)
    ws5.cell(br + 7, 3, len(SEGMENT_ORDER)); ws5.cell(br + 7, 4, total_bud)

    # Стилизация
    _apply_visual_style(
        wb, rfm_next=rfm_next, rfm2_next=rfm2_next,
        ref_wb_bytes=config.ref_wb_bytes, has_weather=has_weather,
    )
    _progress(0.95)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    _progress(1.0)
    return buf
